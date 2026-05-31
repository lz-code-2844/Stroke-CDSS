#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
RAG Build Script - Single Knowledge Base Version
Builds one knowledge base at a time to avoid memory issues
"""

import os
import sys
import pickle
from openpyxl import load_workbook

def build_single_kb(excel_file, collection_name):
    """Build a single knowledge base"""

    print(f"\n{'='*60}")
    print(f"Building: {collection_name}")
    print(f"From: {excel_file}")
    print('='*60)

    # Check file
    if not os.path.exists(excel_file):
        print(f"❌ File not found: {excel_file}")
        return False

    # Create output directory
    output_dir = "knowledge_base/simple_rag"
    os.makedirs(output_dir, exist_ok=True)

    # Load Excel
    print("\n[1/4] Loading Excel...")
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    print(f"  Total rows: {ws.max_row - 1}")

    # Read data
    print("\n[2/4] Reading documents...")
    documents = []
    metadatas = []

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 1000 == 0:
            print(f"  Progress: {row_idx - 1}/{ws.max_row - 1}")

        row_values = [cell.value for cell in ws[row_idx]]
        doc_dict = dict(zip(headers, row_values))

        # Check required fields
        if not doc_dict.get('title') or not doc_dict.get('PMID'):
            continue

        # Build document text
        parts = []
        if doc_dict.get('title'):
            parts.append(str(doc_dict['title']))
        if doc_dict.get('authors'):
            parts.append(str(doc_dict['authors']))
        if doc_dict.get('journal'):
            parts.append(str(doc_dict['journal']))
        if doc_dict.get('abstract'):
            parts.append(str(doc_dict['abstract']))

        doc_text = ' '.join(parts)

        documents.append(doc_text)
        metadatas.append({
            'pmid': str(doc_dict.get('PMID', '')),
            'title': str(doc_dict.get('title', ''))[:200],
            'journal': str(doc_dict.get('journal', ''))[:100],
            'year': str(doc_dict.get('year', ''))
        })

    wb.close()
    print(f"\n  ✓ Loaded {len(documents)} documents")

    # Save as pickle (vectorization deferred to runtime)
    print("\n[3/4] Saving data...")
    output_file = os.path.join(output_dir, f"{collection_name}.pkl")

    data = {
        'documents': documents,
        'metadatas': metadatas,
        'collection_name': collection_name
    }

    with open(output_file, 'wb') as f:
        pickle.dump(data, f)

    print(f"  ✓ Saved to: {output_file}")
    print(f"  ✓ File size: {os.path.getsize(output_file) / 1024 / 1024:.2f} MB")

    print("\n[4/4] Complete!")
    print(f"  ✅ {collection_name}: {len(documents)} documents")

    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python build_single_kb.py <kb_number>")
        print("  1 = thrombectomy")
        print("  2 = thrombolysis")
        print("  3 = imaging_triage")
        print("  4 = imaging_scoring")
        sys.exit(1)

    kb_num = sys.argv[1]

    configs = {
        '1': {
            'file': 'knowledge_base/excel/thrombectomy_db.xlsx',
            'name': 'thrombectomy_literature'
        },
        '2': {
            'file': 'knowledge_base/excel/thrombolysis_db.xlsx',
            'name': 'thrombolysis_literature'
        },
        '3': {
            'file': 'knowledge_base/excel/imaging_triage.xlsx',
            'name': 'imaging_triage_literature'
        },
        '4': {
            'file': 'knowledge_base/excel/imaging_scoring.xlsx',
            'name': 'imaging_scoring_literature'
        }
    }

    if kb_num not in configs:
        print(f"❌ Invalid number: {kb_num}")
        sys.exit(1)

    config = configs[kb_num]
    success = build_single_kb(config['file'], config['name'])

    if success:
        print("\n" + "="*60)
        print("  ✅ Build successful!")
        print("="*60)
    else:
        print("\n❌ Build failed!")
        sys.exit(1)
